import sys
import os
import math
import copy
import json
import shutil
import openpyxl

def calc_proposed_effort(role, task):
    r = role.lower()
    t = task.lower()
    if any(k in t for k in ["logic", "backend", "system", "contract", "auth"]):
        if "developer" in r: return 0.95
        if "lead" in r: return 0.35
        return 0.1
    if any(k in t for k in ["ui", "shop", "screen", "design"]):
        if "ui/ux" in r or "2d" in r: return 0.85
        if "developer" in r: return 0.4
        return 0.2
    if any(k in t for k in ["3d", "model", "boss", "map", "animation"]):
        if "3d" in r or "animator" in r: return 0.95
        if "concept" in r: return 0.7
        return 0.2
    if "developer" in r or "artist" in r: return 0.45
    if "qa" in r or "qc" in r: return 0.4
    if "manager" in r: return 0.2
    return 0.1

def generate_project_files(config_path):
    print("STARTING PM EXCEL GENERATOR (V4 - GROUP CASCADING OVRFLW)...")
    
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
        
    base_dir = os.path.dirname(os.path.abspath(__file__))
    plan_out = os.path.join(base_dir, f"{config['project_name'].replace(' ', '_')}_Plan.xlsx")
    cost_out = os.path.join(base_dir, f"{config['project_name'].replace(' ', '_')}_Cost.xlsx")
    
    print(" -> Instantiating Templates...")
    shutil.copy(os.path.join(base_dir, 'templates', 'Project_Plan_Template.xlsx'), plan_out)
    shutil.copy(os.path.join(base_dir, 'templates', 'Cost_Planning_Template.xlsx'), cost_out)
    
    wb_plan = openpyxl.load_workbook(plan_out, data_only=False)
    ds = wb_plan['3.Detail Schedule']
    res = wb_plan['4.Resource']
    
    # Pre-calculate role efforts & Group capacity
    role_efforts = {}
    for ms in config['milestones']:
        for sp in range(ms['sprints_count']):
            for epic in ms['backlog_epics']:
                for member in config['team']:
                    role = member['role']
                    role_efforts[role] = role_efforts.get(role, 0) + calc_proposed_effort(role, epic)

    role_capacity = {}
    for member in config['team']:
        role = member['role']
        role_capacity[role] = role_capacity.get(role, 0) + float(config["project_duration_months"])
        
    tbd_insertions = {}
    for role, tot in role_efforts.items():
        cap = role_capacity.get(role, 0)
        if tot > cap:
            tbd_insertions[role] = math.ceil((tot - cap) / float(config["project_duration_months"]))

    # 2. Build 4.Resource Dashboard with Spawning
    print(" -> Building Dual-Tracker Resource Dashboard (Cascading)...")
    base_row = 57
    tbd_spawned = set()
    
    for t in config['team']:
        res.cell(row=base_row, column=1).value = t['index']
        res.cell(row=base_row, column=2).value = t['name']
        res.cell(row=base_row, column=3).value = t['role']
        
        res.cell(row=base_row, column=5).value = t['proposed_ra'] # RA THỰC TẾ
        res.cell(row=base_row, column=9).value = f'=E{base_row}*{config["project_duration_months"]}' # EFFORT THỰC TẾ
        res.cell(row=base_row, column=6).value = t['start']
        res.cell(row=base_row, column=7).value = t['end']
        
        res.cell(row=base_row, column=4).number_format = '0%'
        res.cell(row=base_row, column=5).number_format = '0%'
        res.cell(row=base_row, column=8).number_format = '0.00'
        res.cell(row=base_row, column=9).number_format = '0.00'
        
        tbd_count = 0
        if t['role'] in tbd_insertions and t['role'] not in tbd_spawned:
            tbd_count = tbd_insertions[t['role']]
            res.insert_rows(base_row + 1, amount=tbd_count)
            tbd_spawned.add(t['role'])
            
            for t_idx in range(tbd_count):
                t_row = base_row + 1 + t_idx
                for c in range(1,10):
                    res.cell(row=t_row, column=c).font = copy.copy(res.cell(row=base_row, column=c).font)
                    res.cell(row=t_row, column=c).alignment = copy.copy(res.cell(row=base_row, column=c).alignment)
                    res.cell(row=t_row, column=c).border = copy.copy(res.cell(row=base_row, column=c).border)
                    res.cell(row=t_row, column=c).number_format = res.cell(row=base_row, column=c).number_format
                    
                res.cell(row=t_row, column=1).value = f"{t['index']}.{t_idx+1}_TBD"
                res.cell(row=t_row, column=2).value = "TBD"
                res.cell(row=t_row, column=3).value = t['role']
                res.cell(row=t_row, column=5).value = 0 
                res.cell(row=t_row, column=9).value = 0 

        for p_idx in range(base_row, base_row + tbd_count + 1):
            clean_role = t['role'].replace('"', '""')
            sum_logic = f'SUMIFS(\'3.Detail Schedule\'!F:F, \'3.Detail Schedule\'!C:C, "*{clean_role}*")'
            consumed = f'SUMIFS(H$56:H{p_idx-1}, C$56:C{p_idx-1}, "*{clean_role}*")' if p_idx > 57 else "0"
            res.cell(row=p_idx, column=8).value = f'=MIN({config["project_duration_months"]}, MAX(0, {sum_logic} - {consumed}))'
            res.cell(row=p_idx, column=4).value = f'=H{p_idx}/{config["project_duration_months"]}'
                
        base_row += (tbd_count + 1)

    # 3. Build 3.Detail Schedule
    print(" -> Spawning Milestone & Sprint Sequences...")
    cur_row = 18
    for ms_idx, ms in enumerate(config['milestones']):
        ms_wbs = f"2.{ms_idx+1}"
        ds.cell(row=cur_row, column=1).value = ms_wbs
        ds.cell(row=cur_row, column=2).value = ms['name']
        ms_row = cur_row
        sprint_rows = []
        cur_row += 1
        for sp_idx in range(ms['sprints_count']):
            sp_wbs = f"{ms_wbs}.{sp_idx+1}"
            ds.cell(row=cur_row, column=1).value = sp_wbs
            ds.cell(row=cur_row, column=2).value = f"Sprint {sp_idx+1}"
            sp_row = cur_row
            task_rows = []
            cur_row += 1
            for epic in ms['backlog_epics']:
                for member in config['team']:
                    t_wbs = f"{sp_wbs}.{len(task_rows)+1}"
                    ds.cell(row=cur_row, column=1).value = t_wbs
                    ds.cell(row=cur_row, column=2).value = f"[{epic}] - Review & Implement"
                    ds.cell(row=cur_row, column=3).value = member['role']
                    ds.cell(row=cur_row, column=6).value = calc_proposed_effort(member['role'], epic)
                    task_rows.append(cur_row)
                    cur_row += 1
            if task_rows:
                ds.cell(row=sp_row, column=6).value = f"=SUM({','.join([f'F{r}' for r in task_rows])})"
            sprint_rows.append(sp_row)
        if sprint_rows:
            ds.cell(row=ms_row, column=6).value = f"=SUM({','.join([f'F{r}' for r in sprint_rows])})"
            
    wb_plan.save(plan_out)
    print("GENERATION COMPLETE!")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        generate_project_files(sys.argv[1])
