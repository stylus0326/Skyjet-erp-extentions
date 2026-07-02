"""
Excel Formula Recalculation & Verification Tool

Recalculates all formulas in an Excel file using LibreOffice headless mode,
then scans for common formula errors. Returns structured JSON output.

Usage:
    python3 recalc.py <excel_file> [timeout_seconds]

Requirements:
    - LibreOffice installed (brew install --cask libreoffice / apt install libreoffice)
    - openpyxl (pip install openpyxl)
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from soffice_env import get_soffice_env

MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "RecalcModule.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script"
  script:name="RecalcModule" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

EXCEL_ERRORS = [
    "#VALUE!", "#DIV/0!", "#REF!", "#NAME?",
    "#NULL!", "#NUM!", "#N/A",
]


def has_timeout_cmd(cmd_name):
    """Check if a timeout command is available on the system."""
    try:
        subprocess.run(
            [cmd_name, "--version"],
            capture_output=True, timeout=2, check=False,
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def get_macro_dir():
    """Return the LibreOffice macro directory for the current platform."""
    if platform.system() == "Darwin":
        return os.path.expanduser(MACRO_DIR_MACOS)
    return os.path.expanduser(MACRO_DIR_LINUX)


def setup_macro():
    """Install the recalculation macro into LibreOffice if not present."""
    macro_dir = get_macro_dir()
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if os.path.exists(macro_file):
        content = Path(macro_file).read_text()
        if "RecalculateAndSave" in content:
            return True

    if not os.path.exists(macro_dir):
        try:
            subprocess.run(
                ["soffice", "--headless", "--terminate_after_init"],
                capture_output=True, timeout=15, env=get_soffice_env(),
            )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            return False
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except OSError:
        return False


def scan_errors(filename):
    """Scan an Excel file for formula errors. Returns error details dict."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    error_details = {err: [] for err in EXCEL_ERRORS}
    total_errors = 0
    formula_count = 0

    # Scan for errors (read calculated values)
    wb = load_workbook(filename, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            loc = f"{sheet_name}!{cell.coordinate}"
                            error_details[err].append(loc)
                            total_errors += 1
                            break
    wb.close()

    # Count formulas (read formula strings)
    wb_f = load_workbook(filename, data_only=False)
    for sheet_name in wb_f.sheetnames:
        ws = wb_f[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if (cell.value and isinstance(cell.value, str)
                        and cell.value.startswith("=")):
                    formula_count += 1
    wb_f.close()

    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": formula_count,
    }

    if total_errors > 0:
        result["error_summary"] = {}
        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

    return result


def recalc(filename, timeout=30):
    """Recalculate all formulas in an Excel file using LibreOffice."""
    filepath = Path(filename)
    if not filepath.exists():
        return {"error": f"File not found: {filename}"}

    abs_path = str(filepath.absolute())

    if not setup_macro():
        return {"error": "Failed to setup LibreOffice macro. Is LibreOffice installed?"}

    cmd = [
        "soffice", "--headless", "--norestore",
        "vnd.sun.star.script:Standard.RecalcModule.RecalculateAndSave"
        "?language=Basic&location=application",
        abs_path,
    ]

    # Add timeout wrapper
    sys_name = platform.system()
    if sys_name == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif sys_name == "Darwin" and has_timeout_cmd("gtimeout"):
        cmd = ["gtimeout", str(timeout)] + cmd

    result = subprocess.run(
        cmd, capture_output=True, text=True, env=get_soffice_env(),
    )

    # Return code 124 = timeout (acceptable — LO may hang after saving)
    if result.returncode not in (0, 124):
        err_msg = result.stderr.strip() or "Unknown error during recalculation"
        return {"error": f"LibreOffice recalculation failed: {err_msg}"}

    return scan_errors(abs_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 recalc.py <excel_file> [timeout_seconds]")
        print()
        print("Recalculates all formulas in an Excel file using LibreOffice,")
        print("then scans for formula errors.")
        print()
        print("Returns JSON with:")
        print("  status        : 'success' or 'errors_found'")
        print("  total_errors  : Count of formula errors")
        print("  total_formulas: Count of formulas in file")
        print("  error_summary : Breakdown by error type with cell locations")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    output = recalc(filename, timeout)
    print(json.dumps(output, indent=2))


if __name__ == "__main__":
    main()
